from data_structure_git import *
from openpyxl.styles import Font, colors, fills
from openpyxl.workbook import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl import load_workbook
import csv
from pathlib import Path
import os


class TestCase:
    def __init__(self, tca, result, exitcode):
        self._testCaseName = tca
        self._exitCode = exitcode
        self._result = result
        self._module_git = ''


class Module:
    def __init__(self, name):
        self._name = name

        """"List of Test Cases"""
        self._listTestCases = []

    def append_test_case(self, tca):
        self._listTestCases.append(tca)


class ReportData:
    def __init__(self):
        """List of Modules"""
        self._listModules = []
        self.totalTests = ''
        self.skippedTests = ''
        self.totalFailures = ''
        self.nrTotalTest = 0
        self.nrTotalFailures = 0
        self.nrTotalSkipped = 0
        self.nrTotalPass = 0
        self.percentagePass = 0
        self.percentageConf = 0
        self.percentageFail = 0

    def append_module(self, module):
        self._listModules.append(module)

    def append_total_tests(self, total_str):
        self.totalTests += total_str

    def append_skipped_tests(self, skipped_str):
        self.skippedTests += skipped_str

    def append_total_failures(self, failures_str):
        self.totalFailures += failures_str


class Generator:
    pass_str = 'pass'
    fail_str = 'fail'
    skip_str = 'skip'
    conf_str = 'conf'
    total_tests = 'Total Tests'
    skipped_test = 'Total Skipped'
    total_failures = 'Total Failures'
    antet = 'job'
    git_path_with_tcas= 'runtest'
    report_csv = ReportData()
    report_git = ReportDataGit()
    workbook = Workbook()

    @staticmethod
    def git_runtest_extract_data(git_folder):
        """ Extracts the data from the git runner dir which has the modules with testcases """
        project_folder = Path(__file__).parent
        path = os.path.join(project_folder, git_folder)
        run_test_path = os.path.join(path, Generator.git_path_with_tcas)
        for subdir, dirs, files in os.walk(run_test_path):
            for file in files:
                if file == 'Makefile':
                    continue

                module = ModuleGit(file)
                Generator.report_git.append_module(module)
                with open(os.path.join(run_test_path, file), 'r') as text_file:
                    for line in text_file:
                        if line == '' or '#' in line:  #line empty or comment
                            continue
                        else:
                            words = line.split()
                            if len(words) >= 1:
                                tca = TestCaseGit(words[0])
                                module.append_test_case(tca)

    @staticmethod
    def get_module_from_testcasename(tca):
        """ search the testcase name in the git data structure and return the report name, if not found return current tca name """
        for module in Generator.report_git._listModules:
            for testcase in module._listTestCases:
                if testcase._testCaseName in tca._testCaseName:
                    return module._name
        return tca._testCaseName


    @staticmethod
    def file_parser_ltp(report_file):
        """extracts the data from the csv lava report and apppend it on the data structure"""
        with open(report_file) as csvDataFile:
            csv_reader = csv.reader(csvDataFile)
            for row in csv_reader:
                if Generator.antet in row:
                    continue

                Generator.report_csv.nrTotalTest += 1
                result_tca = row[2]
                name_tca = row[11]

                if result_tca == Generator.skip_str:
                    result_tca = Generator.conf_str

                test_case = TestCase(name_tca, result_tca, 'N/A')
                module_name_from_git = Generator.get_module_from_testcasename(test_case)
                test_case._module_git = module_name_from_git
                module = Module(row[1])
                module.append_test_case(test_case)
                Generator.report_csv.append_module(module)

                if Generator.pass_str in result_tca:
                    Generator.report_csv.nrTotalPass += 1
                elif Generator.fail_str in result_tca:
                    Generator.report_csv.nrTotalFailures += 1
                elif Generator.skip_str in result_tca:
                    Generator.report_csv.nrTotalSkipped += 1
                else: # is a conf test case
                    Generator.report_csv.nrTotalSkipped += 1

        Generator.report_csv.percentagePass = round(Generator.report_csv.nrTotalPass * 100 / Generator.report_csv.nrTotalTest, 2)
        Generator.report_csv.percentageConf = round(Generator.report_csv.nrTotalSkipped * 100 / Generator.report_csv.nrTotalTest, 2)
        Generator.report_csv.percentageFail = round(Generator.report_csv.nrTotalFailures * 100 / Generator.report_csv.nrTotalTest, 2)

    @staticmethod
    def list_test_cases():
        """helper method to see the data colected in the data_structure"""
        for module in Generator.report_git._listModules:
            print('Module name: ', module._name)
            for tca in module._listTestCases:
                print('TestCase: ', 'name: ', tca._testCaseName)

    @staticmethod
    def append_data_into_cells(worksheet):
        """append data into the worksheet"""
        current_row = 5
        current_column = 1
        my_red = colors.Color(colors.RED)
        my_fill_red = fills.PatternFill(patternType='solid', fgColor=my_red)
        my_green = colors.Color(colors.GREEN)
        my_fill_green = fills.PatternFill(patternType='solid', fgColor=my_green)
        my_pink = colors.Color(rgb='FF9999')
        my_fill_pink = fills.PatternFill(patternType='solid', fgColor=my_pink)

        for module in Generator.report_csv._listModules:
            for tca in module._listTestCases:
                worksheet.cell(row=current_row, column=current_column).value = tca._module_git
                current_column += 1
                worksheet.cell(row=current_row, column=current_column).value = tca._testCaseName
                current_column += 1
                worksheet.cell(row=current_row, column=current_column).value = tca._result
                if Generator.pass_str in tca._result:
                    worksheet.cell(row=current_row, column=current_column).fill = my_fill_green
                elif Generator.fail_str in tca._result:
                    worksheet.cell(row=current_row, column=current_column).fill = my_fill_red
                else:
                    worksheet.cell(row=current_row, column=current_column).fill = my_fill_pink

                current_column += 3
                worksheet.cell(row=current_row, column=current_column).value = f"=VLOOKUP(A{current_row},'ES6 - LTP Test Results'!A:B,2)"
                current_column += 1
                worksheet.cell(row=current_row, column=current_column).value = f'=IF(AND(C{current_row}<>F{current_row},F{current_row}<>"N/A"),"Different","OK")'

                # current_column += 1
                # worksheet.cell(row=current_row, column=current_column).value = tca._exitCode
                current_column = 1
                current_row += 1

        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Summary:'
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.total_tests
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.report_csv.nrTotalTest
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.skipped_test
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.report_csv.nrTotalSkipped
        current_row += 1
        current_column = 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.total_failures
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.report_csv.nrTotalFailures
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Percentage Pass'
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_green
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.report_csv.percentagePass
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_green
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Percentage Fail'
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_red
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.report_csv.percentageFail
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_red
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Percentage Skipped'
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_pink
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.report_csv.percentageConf
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_pink

        current_row -= 2
        current_column = 1
        # at the end create a chart
        chart = DoughnutChart()
        labels = Reference(worksheet, min_col=current_column, min_row=current_row, max_row=current_row + 2)
        data = Reference(worksheet, min_col=current_column + 1, min_row=current_row, max_row=current_row + 2)

        # worksheet.auto_filter.ref  = 'A5:C1766'
        # worksheet.auto_filter.add_sort_condition('A{0}:A{1}'.format(5, 1766))

        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = "LTP test results"

        chart.style = 10
        worksheet.add_chart(chart, "I8")

    @staticmethod
    def create_ltp_test_report_sheet():
        """ format the excel sheet_0, at the end apppend data """
        # Create a workbook
        sheet_0 = Generator.workbook.active
        sheet_0.title = 'Lava_report'

        bold_font = Font(bold=True, color=colors.DARKYELLOW, size=20)

        sheet_0['A1'].font = bold_font

        sheet_0.merge_cells('A1:D1')

        sheet_0['A1'] = 'LTP Test report'  # + str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        sheet_0['A3'] = 'Module'
        sheet_0['B3'] = 'Test Case'
        sheet_0['C3'] = 'Result'
        sheet_0['E3'] = 'Observations'
        sheet_0['F3'] = 'ES6 - LTP Results'
        sheet_0['G3'] = 'Results comparison'

        # set the width of the column
        sheet_0.column_dimensions['A'].width = 30
        sheet_0.column_dimensions['B'].width = 50
        sheet_0.column_dimensions['C'].width = 10
        sheet_0.column_dimensions['E'].width = 20
        sheet_0.column_dimensions['F'].width = 20
        sheet_0.column_dimensions['G'].width = 20

        Generator.append_data_into_cells(sheet_0)


    @staticmethod
    def create_es6_sheet(filepath):
        """ format the excel sheet_0, at the end apppend data """
        # Create a workbook
        sheet_1 = Generator.workbook.create_sheet('Sheet_B')
        sheet_1.title = 'ES6 - LTP Test Results'

        # sheet_1['A1'] = 'Test Case Name'
        # sheet_1['B1'] = 'Result'
        # sheet_1['C1'] = 'Exit Value'
        sheet_1.column_dimensions['A'].width = 30
        sheet_1.column_dimensions['B'].width = 20
        sheet_1.column_dimensions['C'].width = 10

        wb_es6 = load_workbook(filepath)
        source = wb_es6.get_sheet_by_name('LTP_Test_Results')    # add a try catch

        counter = 0
        new_rows = []

        for rrow in source.iter_rows():
            new_rows.append([])
            for cell in rrow:
                new_rows[counter].append(cell.value)
            counter += 1
        for wrow in new_rows:
            sheet_1.append(wrow)


    @staticmethod
    def create_lava_job_sheet(filepath):
        """ format the excel sheet_0, at the end apppend data """
        # Create a workbook
        sheet_2 = Generator.workbook.create_sheet('Sheet_C')
        sheet_2.title = 'Lava raw results'

        with open(filepath) as file_csv:
            reader = csv.reader(file_csv, delimiter=',')
            for row in reader:
                sheet_2.append(row)


    @staticmethod
    def save_xcel():
        """ format the excel sheet_0 """
        try:
            Generator.workbook.save(filename="l4b-software___testReport.xlsx")
        except PermissionError as e:
            print("\n\n\n Excel file is open. Please close the excel file !!!")